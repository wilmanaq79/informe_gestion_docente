# -*- coding: utf-8 -*-
"""Agente de verificacion de firmas en los documentos que un docente sube
en la sesion "Entrega de documentos". Analiza cada archivo apenas se
sube y devuelve un veredicto que ayuda al Director, Secretario
Academico y Secretaria del Programa a decidir si aprueban o rechazan la
entrega -- NO reemplaza su criterio, sobre todo para firmas manuscritas
escaneadas, que no se pueden verificar de forma confiable por software.

Niveles de deteccion (PDF y Excel se analizan con la MISMA logica de
fondo, solo cambia de donde se extraen las "lineas" de texto), de mas a
menos confiable:
  1. Firma digital (certificado) embebida en el archivo -- alta confianza.
     PDF: campo de formulario tipo /Sig. Excel (.xlsx): carpeta
     "_xmlsignatures" dentro del paquete OOXML.
  2. El documento tiene, en una misma zona (una linea de PDF o una fila
     de Excel, junto con la linea/fila siguiente), una de estas DOS
     anclas junto al nombre COMPLETO del docente (al menos 2 partes de
     su nombre, no solo una) -- confianza media:
       a) Contexto "firma": el texto menciona "firma"/"firmado"/etc.
          (p.ej. "Firma del docente: Wilman Andres Quiñonez").
       b) Contexto "docente": el texto trae un campo tipo
          "Docente: ____" (sin la palabra "firma" en ningun lado) y ese
          espacio esta completado con el nombre del docente -- ese
          campo ES el punto de firma/responsable en este tipo de
          documento. Solo se usa si el documento NO tiene, en ninguna
          parte, un renglon de firma explicito (ver mas abajo).
     Exigir 2 partes del nombre (no 1) y acotar la busqueda a la
     linea/fila donde aparece la ancla evita falsos positivos: un
     nombre de pila comun del docente (p.ej. "Andres") puede coincidir
     por casualidad con un estudiante en una lista de asistencia; eso
     ya NO alcanza para marcar el documento como firmado.
  3. El archivo trae imagenes incrustadas (posible firma/sello
     escaneado) pero nada de lo anterior -- indeterminado, requiere
     revision humana.
  4. Nada de lo anterior -- probablemente no firmado.
  5. Imagen suelta (jpg/png) -- no se puede verificar automaticamente
     una firma manuscrita en una imagen -- indeterminado, requiere
     revision humana.

Prioridad entre las anclas 2a y 2b (bug real corregido): si el
documento tiene algun renglon con "firma"/"firmado"/etc en cualquier
parte, ESE renglon manda y el ancla 2b ("Docente: ____") deja de
consultarse -- si ninguno de esos renglones de firma trae el nombre
completo, el documento se da por NO firmado. Antes de esta correccion,
un reporte de Academusoft con un encabezado "Identificación Docente:
NOMBRE" (mera identificacion, no firma) se marcaba como firmado aunque
el renglon real "Firma Del Docente: ______" mas abajo estuviera en
blanco, porque el ancla 2b encontraba el encabezado antes de llegar al
renglon de firma vacio.
"""
import re
import unicodedata
import zipfile
from io import BytesIO

import pdfplumber
from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader

PALABRAS_FIRMA = ("firma", "firmado", "firmó", "firmo", "suscrito", "suscribe", "suscrita")

# Una imagen mas pequeña que esto en ambas dimensiones es, con altisima
# probabilidad, un icono/logo decorativo de membrete (p.ej. el escudo de
# la universidad en un reporte de Academusoft) y NO una firma o sello
# escaneado -- no debe disparar el veredicto "indeterminado, requiere
# revision manual". Una firma pegada como imagen real observada en
# pruebas mide 169x74 px, muy por encima de este umbral.
TAMANO_MINIMO_IMAGEN_FIRMA = 40


def _normalizar(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto if not unicodedata.combining(c)).lower()


def _contiene_palabra(texto_normalizado: str, palabra: str) -> bool:
    """Coincidencia de palabra completa (no substring): "docente" no debe
    disparar dentro de otra palabra, y "firma" no debe disparar dentro de
    "confirma" o "confirmado"."""
    return re.search(rf"\b{re.escape(palabra)}\b", texto_normalizado) is not None


def _nombre_coincide_fuerte(texto_normalizado: str, nombre_docente: str) -> bool:
    """Exige al menos 2 partes distintas del nombre del docente (nombres y
    apellidos de mas de 3 letras) si hay 2 o mas disponibles; con un
    nombre de una sola palabra, exige esa unica parte. Esto evita que un
    nombre de pila comun (p.ej. "Andres") haga match por casualidad con
    otra persona mencionada en el mismo documento."""
    partes = [p for p in _normalizar(nombre_docente).split() if len(p) > 3]
    if not partes:
        return False
    coincidencias = sum(1 for p in partes if _contiene_palabra(texto_normalizado, p))
    if len(partes) == 1:
        return coincidencias == 1
    return coincidencias >= 2


def _buscar_firma_en_lineas(lineas: list, nombre_docente: str) -> str | None:
    """Recorre las lineas (PDF) o filas (Excel) de un documento buscando
    evidencia de firma. Revisa cada linea junto con la siguiente, por si
    la etiqueta ("Firma:", "Docente:") y el nombre completado quedan en
    lineas/filas separadas. Devuelve 'firma' o 'docente' segun cual
    ancla disparo el hallazgo, o None si no se encontro evidencia.

    Si el documento trae un campo de firma explicito ("Firma"/
    "Firmado"/etc.), ESE campo manda sobre el ancla mas debil
    "Docente: ____", y se revisa PRIMERO: muchos reportes
    institucionales (p.ej. Academusoft) traen un encabezado de mera
    identificacion ("Identificación Docente: NOMBRE") que no es una
    firma, seguido mas abajo por el verdadero renglon de firma en
    blanco ("Firma Del Docente: ________"). Si se revisara el ancla
    "docente" primero, ese encabezado de identificacion se confundia
    con una firma real (bug reportado: PDF de Academusoft marcado como
    firmado sin estarlo). Por eso: si existe algun renglon de firma en
    el documento, solo esos renglones deciden el resultado -- si
    ninguno trae el nombre completo, el documento se da por NO firmado,
    sin recurrir al encabezado de identificacion como respaldo."""
    lineas_norm = [_normalizar(l) for l in lineas]
    ventanas = []
    for i, actual in enumerate(lineas_norm):
        siguiente = lineas_norm[i + 1] if i + 1 < len(lineas_norm) else ""
        ventanas.append(f"{actual} {siguiente}")

    ventanas_firma = [v for v in ventanas if any(_contiene_palabra(v, palabra) for palabra in PALABRAS_FIRMA)]
    if ventanas_firma:
        for ventana in ventanas_firma:
            if _nombre_coincide_fuerte(ventana, nombre_docente):
                return "firma"
        return None

    for ventana in ventanas:
        if _contiene_palabra(ventana, "docente") and _nombre_coincide_fuerte(ventana, nombre_docente):
            return "docente"

    return None


_DETALLE_HALLAZGO = {
    "firma": "El documento menciona 'firma' junto al nombre completo del docente.",
    "docente": "El documento tiene un campo 'Docente' completado con el nombre completo del docente (punto de firma/responsable de este tipo de documento).",
}


# --- PDF ----------------------------------------------------------------

def _tiene_firma_digital_pdf(contenido: bytes) -> bool:
    try:
        reader = PdfReader(BytesIO(contenido))
        campos = reader.get_fields() or {}
        return any(campo.get("/FT") == "/Sig" for campo in campos.values())
    except Exception:
        return False


def _lineas_pdf(contenido: bytes) -> list:
    try:
        with pdfplumber.open(BytesIO(contenido)) as pdf:
            lineas = []
            for pagina in pdf.pages:
                texto = pagina.extract_text() or ""
                lineas.extend(texto.split("\n"))
            return lineas
    except Exception:
        return []


def _tiene_imagenes_pdf(contenido: bytes) -> bool:
    """Ignora imagenes mas pequeñas que TAMANO_MINIMO_IMAGEN_FIRMA en
    ambas dimensiones (iconos/logos de membrete) -- solo cuentan como
    "posible firma o sello" las imagenes de tamaño creible para eso."""
    try:
        reader = PdfReader(BytesIO(contenido))
        for pagina in reader.pages:
            recursos = pagina.get("/Resources") or {}
            xobjects = recursos.get("/XObject")
            if not xobjects:
                continue
            for objeto in xobjects.values():
                img = objeto.get_object()
                if img.get("/Subtype") != "/Image":
                    continue
                ancho, alto = img.get("/Width", 0), img.get("/Height", 0)
                if ancho >= TAMANO_MINIMO_IMAGEN_FIRMA or alto >= TAMANO_MINIMO_IMAGEN_FIRMA:
                    return True
    except Exception:
        pass
    return False


def _analizar_pdf(contenido: bytes, nombre_docente: str) -> dict:
    if _tiene_firma_digital_pdf(contenido):
        return {
            "firma_detectada": True,
            "confianza": "alta",
            "detalle": "Firma digital (certificado) detectada en el PDF.",
        }

    hallazgo = _buscar_firma_en_lineas(_lineas_pdf(contenido), nombre_docente)
    if hallazgo:
        return {
            "firma_detectada": True,
            "confianza": "media",
            "detalle": _DETALLE_HALLAZGO[hallazgo],
        }

    if _tiene_imagenes_pdf(contenido):
        return {
            "firma_detectada": None,
            "confianza": "baja",
            "detalle": (
                "El PDF contiene imágenes (posible firma o sello escaneado) — no se puede confirmar "
                "automáticamente una firma manuscrita, requiere revisión manual."
            ),
        }

    return {
        "firma_detectada": False,
        "confianza": "media",
        "detalle": (
            "No se detectó firma digital, ni un campo de 'firma' o 'Docente' completado con el nombre "
            "del docente, ni imágenes en el documento."
        ),
    }


# --- Excel (.xlsx) --------------------------------------------------------
# Un .xlsx es en realidad un archivo .zip: se puede inspeccionar su
# contenido interno sin depender de que openpyxl "entienda" cada detalle.

def _tiene_firma_digital_xlsx(contenido: bytes) -> bool:
    try:
        with zipfile.ZipFile(BytesIO(contenido)) as z:
            return any(nombre.startswith("_xmlsignatures") for nombre in z.namelist())
    except Exception:
        return False


def _tiene_imagenes_xlsx(contenido: bytes) -> bool:
    """Igual que _tiene_imagenes_pdf: ignora imagenes mas pequeñas que
    TAMANO_MINIMO_IMAGEN_FIRMA en ambas dimensiones (iconos/logos de
    membrete pegados en una celda) -- solo cuentan las de tamaño
    creible para ser una firma o sello escaneado."""
    try:
        with zipfile.ZipFile(BytesIO(contenido)) as z:
            medios = [nombre for nombre in z.namelist() if nombre.startswith("xl/media/")]
            for nombre in medios:
                try:
                    with Image.open(BytesIO(z.read(nombre))) as img:
                        ancho, alto = img.size
                except Exception:
                    continue
                if ancho >= TAMANO_MINIMO_IMAGEN_FIRMA or alto >= TAMANO_MINIMO_IMAGEN_FIRMA:
                    return True
    except Exception:
        pass
    return False


def _lineas_xlsx(contenido: bytes) -> list:
    """Una "linea" aqui es el texto de una fila completa (todas sus
    celdas y comentarios concatenados), para poder acotar la busqueda de
    firma a la fila donde aparece la ancla en vez de a la hoja entera."""
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True)
    except Exception:
        return []

    lineas = []
    try:
        for hoja in wb.worksheets:
            for fila in hoja.iter_rows():
                partes = []
                for celda in fila:
                    if celda.value is not None:
                        partes.append(str(celda.value))
                    if celda.comment is not None and celda.comment.text:
                        partes.append(celda.comment.text)
                if partes:
                    lineas.append(" ".join(partes))
    finally:
        wb.close()
    return lineas


def _analizar_xlsx(contenido: bytes, nombre_docente: str) -> dict:
    if _tiene_firma_digital_xlsx(contenido):
        return {
            "firma_detectada": True,
            "confianza": "alta",
            "detalle": "Firma digital (certificado) detectada en el archivo de Excel.",
        }

    hallazgo = _buscar_firma_en_lineas(_lineas_xlsx(contenido), nombre_docente)
    if hallazgo:
        return {
            "firma_detectada": True,
            "confianza": "media",
            "detalle": _DETALLE_HALLAZGO[hallazgo],
        }

    if _tiene_imagenes_xlsx(contenido):
        return {
            "firma_detectada": None,
            "confianza": "baja",
            "detalle": (
                "El Excel tiene una imagen incrustada (posible firma o sello pegado en una celda) — "
                "no se puede confirmar automáticamente una firma manuscrita, requiere revisión manual."
            ),
        }

    return {
        "firma_detectada": False,
        "confianza": "media",
        "detalle": (
            "No se encontró ningún campo de 'firma' o 'Docente' completado con el nombre del docente, "
            "ni una imagen incrustada, en el archivo de Excel. Si el docente firmó en una hoja aparte, "
            "adjúntala también o pega la firma escaneada en una celda del archivo."
        ),
    }


# --- Punto de entrada ------------------------------------------------------

def analizar_documento(nombre_archivo: str, contenido: bytes, nombre_docente: str) -> dict:
    """Devuelve {'firma_detectada': True/False/None, 'confianza':
    'alta'/'media'/'baja', 'detalle': str}. None en firma_detectada
    significa indeterminado -- ni confirmado ni descartado, requiere que
    un humano lo revise (típicamente firmas manuscritas escaneadas)."""
    extension = nombre_archivo.lower().rsplit(".", 1)[-1] if "." in nombre_archivo else ""

    if extension == "pdf":
        return _analizar_pdf(contenido, nombre_docente)

    if extension == "xlsx":
        return _analizar_xlsx(contenido, nombre_docente)

    if extension in ("jpg", "jpeg", "png"):
        return {
            "firma_detectada": None,
            "confianza": "baja",
            "detalle": "Es una imagen — no se puede confirmar automáticamente una firma manuscrita, requiere revisión manual.",
        }

    return {
        "firma_detectada": None,
        "confianza": "baja",
        "detalle": f"Tipo de archivo '.{extension}' no soportado para verificación automática de firma.",
    }


def resumen_entrega(documentos) -> dict:
    """documentos: lista de objetos con .firma_detectada (bool|None),
    .tipo_documento (str) y .nombre_archivo (str) -- normalmente
    entrega.documentos. Devuelve un resumen agregado para mostrar de un
    vistazo si la entrega completa cumple con las firmas."""
    pendientes = [d for d in documentos if d.firma_detectada is not True]
    return {
        "todos_firmados": len(documentos) > 0 and len(pendientes) == 0,
        "documentos_pendientes": pendientes,
    }
