"""
Logica central del agente de llenado de notas para el
'FORMATO GESTION Y AUTOEVALUACION DOCENTE' (MI-DO-FO16).

Este modulo es compartido por:
  - agente_llenado_notas.py  (linea de comandos)
  - app.py                   (formulario web con Streamlit)

Reglas de negocio (ver pestana INSTRUCTIVO del formato):
  - Matriculados       : numero de estudiantes en el PDF de notas.
  - Asistencia regular : estudiantes con 0 faltas en el corte (100% de
                          asistencia). No viene en el PDF de notas de
                          Academusoft, se toma de una planilla de asistencia
                          por semanas.
  - Evaluados          : estudiantes con al menos una nota registrada en
                          los cortes ya corridos (segun el corte elegido en
                          el formulario).
  - Aprobaron          : nota definitiva >= 60.
        * Corte 1 o Corte 2 (parcial): se ESTIMA proyectando el acumulado
          de los cortes ya corridos sobre 100 puntos, porque a�n falta
          peso por calificar. Se marca con un comentario en la celda.
        * Corte 3 / Final: calculo exacto con los 3 cortes.
  - Inasistencia, Reprobados y los dos porcentajes YA son formulas en el
    Excel: este modulo nunca las toca, solo escribe los 4 valores de arriba.
"""
import re
import shutil
import unicodedata

import pdfplumber
from openpyxl import load_workbook
from openpyxl.comments import Comment

UMBRAL_APROBACION = 60.0

PESO_CORTE = {1: 0.3, 2: 0.3, 3: 0.4}

FILA_OFFSET = {
    "matriculados": 1,
    "asistencia_regular": 2,
    "evaluados": 4,
    "aprobaron": 5,
}

FILA_PATRON = re.compile(
    r"^(?P<no>\d+)\s+"
    r"(?P<tipo_doc>[A-Z]{2})\s*-\s*(?P<doc>\d+)\s+"
    r"(?P<nombre>.+?)\s+"
    r"(?P<c1>[\d.]+)\.\s+"
    r"(?P<c2>[\d.]+)\.\s+"
    r"(?P<c3>-|[\d.]+\.?)\s+"
    r"(?P<extra>-|[\d.]+\.?)\s+"
    r"(?P<pond>[\d.]+)\.\s+"
    r"(?P<repitente>SI|NO)\s*$"
)


def normalizar(texto):
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.upper().strip()


def leer_pdf_notas(pdf_file):
    """Extrae materia y notas por estudiante desde el PDF de Academusoft.

    pdf_file puede ser una ruta (str/Path) o un objeto tipo archivo
    (p.ej. el resultado de st.file_uploader en Streamlit).
    """
    with pdfplumber.open(pdf_file) as pdf:
        texto = "\n".join(page.extract_text() for page in pdf.pages)

    materia_match = re.search(r"([A-Z]{2}\d{4}-[^\n]+)\s+([A-Z0-9]+-[^\n]+)", texto)
    materia = materia_match.group(1).split("-", 1)[1].strip() if materia_match else None
    grupo_match = re.search(r"[A-Z]{2}\d{4}-[^\n]+\s+([A-Z0-9]+-[^\n]+)", texto)
    grupo = grupo_match.group(1).strip() if grupo_match else None

    estudiantes = []
    for linea in texto.splitlines():
        m = FILA_PATRON.match(linea.strip())
        if not m:
            continue
        d = m.groupdict()
        estudiantes.append(
            {
                "no": int(d["no"]),
                "documento": f"{d['tipo_doc']} - {d['doc']}",
                "nombre": d["nombre"],
                "corte1": float(d["c1"]),
                "corte2": float(d["c2"]),
                "corte3": None if d["c3"] == "-" else float(d["c3"].rstrip(".")),
                "pond": float(d["pond"]),
                "repitente": d["repitente"] == "SI",
            }
        )

    if not estudiantes:
        raise ValueError(
            "No se reconocio ninguna fila de estudiante en el PDF. "
            "Verifica que sea un reporte 'Ver Calificaciones' de Academusoft."
        )

    return materia, grupo, estudiantes


def leer_asistencia_excel(path_or_file):
    """Lee una planilla de asistencia por semanas (columnas 'Semana N', P/A)
    y devuelve cuantos estudiantes asistieron al 100% de las semanas del
    corte (Asistencia regular, segun el INSTRUCTIVO del formato)."""
    wb = load_workbook(path_or_file, data_only=False)
    ws = wb.worksheets[0]

    header_row = None
    semana_cols = []
    no_col = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        etiquetas = {}
        for c in row:
            if c.value is not None:
                etiquetas.setdefault(str(c.value).strip().lower(), c.column)
        if any(k.startswith("semana") for k in etiquetas):
            header_row = row[0].row
            semana_cols = [col for etiqueta, col in etiquetas.items() if etiqueta.startswith("semana")]
            no_col = etiquetas.get("no") or etiquetas.get("no.")
            break

    if header_row is None:
        wb.close()
        raise ValueError("No se encontraron columnas 'Semana N' en la planilla de asistencia.")
    if no_col is None:
        wb.close()
        raise ValueError("No se encontro una columna 'No' (numero de estudiante) en la planilla de asistencia.")

    total = 0
    regulares = 0
    for r in range(header_row + 1, ws.max_row + 1):
        # El bloque de estudiantes termina donde la columna 'No' deja de ser
        # un numero (lo siguiente son filas de resumen, no mas estudiantes).
        if not isinstance(ws.cell(row=r, column=no_col).value, (int, float)):
            break
        marcas = [ws.cell(row=r, column=col).value for col in semana_cols]
        marcas = [str(m).strip().upper() for m in marcas if m is not None and str(m).strip() != ""]
        total += 1
        faltas = sum(1 for m in marcas if m == "A")
        if faltas == 0:
            regulares += 1

    wb.close()
    return {"matriculados_asistencia": total, "asistencia_regular": regulares}


def definitiva_estudiante(estudiante, corte):
    """Nota definitiva (0-100) de un estudiante segun el corte disponible.
    Corte 3: calculo exacto con los 3 cortes. Corte 1 o 2: proyeccion del
    acumulado sobre 100 puntos (usada tanto para Aprobados como para el
    dashboard de rendimiento)."""
    if corte == 3:
        if estudiante["corte3"] is None:
            raise ValueError(
                f"Seleccionaste 'Corte 3 / Final' pero el estudiante "
                f"{estudiante['nombre']} no tiene nota de Corte 3 en el PDF. "
                "Verifica que el PDF cargado sea el del cierre del curso."
            )
        return estudiante["corte1"] * 0.3 + estudiante["corte2"] * 0.3 + estudiante["corte3"] * 0.4
    if corte == 2:
        # Proyeccion: acumulado (60% del peso) reescalado sobre 100.
        return estudiante["pond"] / 0.6 if estudiante["pond"] else 0
    # corte == 1: proyeccion, asume que el desempeno de Corte 1 se mantendria.
    return estudiante["corte1"]


def acumulado_ponderado(estudiante, corte):
    """Def. Pond: acumulado ponderado real hasta el corte indicado, con los
    pesos del acuerdo pedagogico (Corte 1 = 30%, Corte 2 = 30%, Corte 3 =
    40%). A diferencia de definitiva_estudiante() esto NO es una proyeccion
    reescalada a 100 -- es la suma de puntos que el estudiante ya tiene
    asegurados sobre el total de 100."""
    acumulado = estudiante["corte1"] * PESO_CORTE[1]
    if corte >= 2:
        acumulado += estudiante["corte2"] * PESO_CORTE[2]
    if corte >= 3:
        if estudiante["corte3"] is None:
            raise ValueError(
                f"El estudiante {estudiante['nombre']} no tiene nota de Corte 3 en el PDF."
            )
        acumulado += estudiante["corte3"] * PESO_CORTE[3]
    return acumulado


def analizar_progreso(estudiante, corte):
    """Progreso de un estudiante hacia la aprobacion (>=60 puntos sobre
    100), corte a corte, usando el acumulado ponderado real (Def. Pond).

    Devuelve un dict con:
      pond           -- acumulado ponderado hasta este corte (Def. Pond)
      peso_restante  -- fraccion del 100% que todavia falta por calificar
      necesaria      -- nota (0-100) que el estudiante necesita sacar en lo
                         que falta para llegar exactamente a 60. None si ya
                         se jugo todo (corte 3).
      estado         -- 'aprobado' / 'reprobado' (corte 3, ya definitivo) o
                         'asegurado' / 'en_riesgo' / 'matematicamente_reprobado'
                         (corte 1 o 2, todavia falta nota por calificar)
    """
    peso_corrido = sum(PESO_CORTE[c] for c in range(1, corte + 1))
    peso_restante = round(1.0 - peso_corrido, 4)
    pond = acumulado_ponderado(estudiante, corte)

    if corte == 3:
        return {
            "pond": pond,
            "peso_restante": 0.0,
            "necesaria": None,
            "estado": "aprobado" if pond >= UMBRAL_APROBACION else "reprobado",
        }

    faltante = UMBRAL_APROBACION - pond
    if faltante <= 0:
        return {"pond": pond, "peso_restante": peso_restante, "necesaria": 0.0, "estado": "asegurado"}

    necesaria = faltante / peso_restante
    estado = "matematicamente_reprobado" if necesaria > 100 else "en_riesgo"
    return {"pond": pond, "peso_restante": peso_restante, "necesaria": necesaria, "estado": estado}


def calcular_resumen(estudiantes, corte, asistencia_regular=None):
    """corte: 1, 2 o 3 (3 = Corte 3 / Final). Determina hasta que corte se
    consideran las notas para Evaluados/Aprobados."""
    if corte not in (1, 2, 3):
        raise ValueError("corte debe ser 1, 2 o 3")

    matriculados = len(estudiantes)
    campos_corte = ["corte1", "corte2", "corte3"][:corte]

    evaluados = sum(
        1 for e in estudiantes if any((e[c] or 0) > 0 for c in campos_corte)
    )

    es_estimado = corte < 3
    aprobaron = sum(
        1 for e in estudiantes if definitiva_estudiante(e, corte) >= UMBRAL_APROBACION
    )

    return {
        "corte": corte,
        "matriculados": matriculados,
        "asistencia_regular": asistencia_regular,
        "evaluados": evaluados,
        "aprobaron": aprobaron,
        "es_estimado": es_estimado,
    }


def localizar_bloque_materia(ws, materia_buscada):
    """Busca en la columna B la fila donde empieza el bloque de la materia."""
    objetivo = normalizar(materia_buscada)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]
        if cell.value and normalizar(str(cell.value)) == objetivo:
            return cell.row
    # Coincidencia parcial si no hay match exacto
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]
        if cell.value and objetivo in normalizar(str(cell.value)):
            return cell.row
    return None


def listar_materias(excel_path_or_file):
    """Devuelve la lista de nombres de materia disponibles en la columna B
    del bloque de docencia (filas 26 a 79 de 'INFORME FINAL')."""
    wb = load_workbook(excel_path_or_file, data_only=False)
    ws = wb["INFORME FINAL"]
    materias = []
    for row in ws.iter_rows(min_row=26, max_row=79, min_col=2, max_col=2):
        cell = row[0]
        if cell.value and str(cell.value).strip():
            materias.append(str(cell.value).strip())
    wb.close()
    return materias


def abrir_plantilla(excel_path_or_file, out_path):
    """Copia la plantilla a out_path y la abre como Workbook, lista para que
    escribir_bloque() escriba una o varias materias antes de guardar()."""
    if hasattr(excel_path_or_file, "read"):
        excel_path_or_file.seek(0)
        with open(out_path, "wb") as f:
            shutil.copyfileobj(excel_path_or_file, f)
    else:
        shutil.copy(excel_path_or_file, out_path)
    return load_workbook(out_path, data_only=False)


def escribir_bloque(wb, materia, resumen):
    """Escribe los 4 valores calculados en el bloque de una materia dentro
    de un Workbook ya abierto (no guarda; permite llamar varias veces para
    varias materias antes de guardar una sola vez)."""
    ws = wb["INFORME FINAL"]

    fila_base = localizar_bloque_materia(ws, materia)
    if fila_base is None:
        raise ValueError(
            f"No se encontro la materia '{materia}' en la columna B de 'INFORME FINAL'. "
            "Revisa que el nombre coincida con el usado en el formato."
        )

    def celda_L(offset):
        return ws.cell(row=fila_base + offset, column=12)  # columna L

    celda_L(FILA_OFFSET["matriculados"]).value = resumen["matriculados"]
    celda_L(FILA_OFFSET["evaluados"]).value = resumen["evaluados"]

    aprobados_cell = celda_L(FILA_OFFSET["aprobaron"])
    aprobados_cell.value = resumen["aprobaron"]
    aprobados_cell.comment = (
        Comment(
            f"Estimado a partir del acumulado hasta Corte {resumen['corte']} "
            "(aun faltan cortes por calificar). Verificar cuando se tenga la "
            "nota definitiva.",
            "Agente de llenado de notas",
        )
        if resumen["es_estimado"]
        else None
    )

    if resumen["asistencia_regular"] is not None:
        celda_L(FILA_OFFSET["asistencia_regular"]).value = resumen["asistencia_regular"]
    else:
        celda_L(FILA_OFFSET["asistencia_regular"]).comment = Comment(
            "El PDF de notas no trae asistencia. Completar manualmente con el "
            "control de asistencia fisico/plataforma, o cargar la planilla de "
            "asistencia en el formulario.",
            "Agente de llenado de notas",
        )

    return fila_base


def guardar(wb, out_path):
    wb.save(out_path)
    wb.close()


def escribir_en_excel(excel_path_or_file, out_path, materia, resumen):
    """Atajo para una sola materia: abre, escribe y guarda de una vez."""
    wb = abrir_plantilla(excel_path_or_file, out_path)
    try:
        fila_base = escribir_bloque(wb, materia, resumen)
    except ValueError:
        wb.close()
        raise
    guardar(wb, out_path)
    return fila_base
