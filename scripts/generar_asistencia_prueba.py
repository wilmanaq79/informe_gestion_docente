"""Genera una planilla de asistencia de PRUEBA para Corte 2 (5 semanas)
de SISTEMAS OPERATIVOS, con los mismos 31 estudiantes del PDF de notas.
Dos estudiantes quedan con 2 faltas cada uno; el resto asiste a las 5 semanas.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

RAIZ = Path(__file__).resolve().parent.parent
PDF = RAIZ / "ejemplos" / "Notas_sistemas_opertivo_corte_2.pdf"
OUT = RAIZ / "ejemplos" / "Asistencia_sistemas_operativo_corte_2_PRUEBA.xlsx"

PAT = re.compile(
    r"^(?P<no>\d+)\s+(?P<tipo>[A-Z]{2})\s*-\s*(?P<doc>\d+)\s+(?P<nombre>.+?)\s+"
    r"(?P<c1>[\d.]+)\.\s+(?P<c2>[\d.]+)\.\s+(?P<c3>-|[\d.]+\.?)\s+(?P<extra>-|[\d.]+\.?)\s+"
    r"(?P<pond>[\d.]+)\.\s+(?P<rep>SI|NO)\s*$"
)

with pdfplumber.open(PDF) as pdf:
    texto = "\n".join(p.extract_text() for p in pdf.pages)

estudiantes = []
for linea in texto.splitlines():
    m = PAT.match(linea.strip())
    if m:
        d = m.groupdict()
        estudiantes.append({"no": int(d["no"]), "doc": f"{d['tipo']} - {d['doc']}", "nombre": d["nombre"]})

# Dos estudiantes con 2 faltas cada uno (semanas distintas), el resto asiste completo.
FALTAS = {
    9: {2, 4},    # CACERES MOSQUERA BRAYAN STIVEN -> falta semana 2 y 4
    11: {1, 3},   # CANDELO GARCES BRAYAN ESTEBAN -> falta semana 1 y 3
}

wb = Workbook()
ws = wb.active
ws.title = "Asistencia Corte 2"

bold = Font(bold=True)
header_fill = PatternFill("solid", fgColor="D9E1F2")
amarillo = PatternFill("solid", fgColor="FFFF00")

ws.merge_cells("A1:L1")
ws["A1"] = "CONTROL DE ASISTENCIA - CORTE 2 (Semanas 1 a 5) - SISTEMAS OPERATIVOS - Grupo IS05D2"
ws["A1"].font = Font(bold=True, size=12)

ws.append([])
headers = ["No", "Documento", "Nombre", "Semana 1", "Semana 2", "Semana 3", "Semana 4", "Semana 5", "Faltas", "Asistencia %", "Asistencia Regular"]
ws.append(headers)
header_row = ws.max_row
for col in range(1, len(headers) + 1):
    c = ws.cell(row=header_row, column=col)
    c.font = bold
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", wrap_text=True)

ws.cell(row=header_row, column=9).comment = Comment(
    "Numero de semanas del corte (de las 5) en que el estudiante NO asistio.", "Agente"
)
ws.cell(row=header_row, column=11).comment = Comment(
    "SI solo si el estudiante asistio al 100% de las clases del corte (0 faltas), "
    "segun la definicion del INSTRUCTIVO del formato de gestion docente.", "Agente"
)

first_data_row = header_row + 1
for i, e in enumerate(estudiantes):
    r = first_data_row + i
    ws.cell(row=r, column=1, value=e["no"])
    ws.cell(row=r, column=2, value=e["doc"])
    ws.cell(row=r, column=3, value=e["nombre"])
    faltas_semanas = FALTAS.get(e["no"], set())
    for semana in range(1, 6):
        col = 3 + semana
        val = "A" if semana in faltas_semanas else "P"
        cell = ws.cell(row=r, column=col, value=val)
        cell.alignment = Alignment(horizontal="center")
        if val == "A":
            cell.fill = amarillo
    ws.cell(row=r, column=9, value=f"=COUNTIF(D{r}:H{r},\"A\")")
    ws.cell(row=r, column=10, value=f"=IFERROR((5-I{r})/5,\"\")")
    ws.cell(row=r, column=10).number_format = "0%"
    ws.cell(row=r, column=11, value=f'=IF(I{r}=0,"SI","NO")')

last_data_row = first_data_row + len(estudiantes) - 1

ws.append([])
resumen_row = ws.max_row + 1
ws.cell(row=resumen_row, column=3, value="Matriculados").font = bold
ws.cell(row=resumen_row, column=4, value=f"=COUNTA(C{first_data_row}:C{last_data_row})")
ws.cell(row=resumen_row + 1, column=3, value="Asistencia regular (100% asistencia)").font = bold
ws.cell(row=resumen_row + 1, column=4, value=f'=COUNTIF(K{first_data_row}:K{last_data_row},"SI")')
ws.cell(row=resumen_row + 2, column=3, value="Inasistencia").font = bold
ws.cell(row=resumen_row + 2, column=4, value=f"=D{resumen_row}-D{resumen_row+1}")

for col, width in zip(range(1, 12), [5, 14, 32, 10, 10, 10, 10, 10, 9, 12, 16]):
    ws.column_dimensions[chr(64 + col)].width = width

wb.save(OUT)
print("Generado:", OUT)
print("Estudiantes:", len(estudiantes))
print("Con 2 faltas:", {e["nombre"]: sorted(FALTAS[e["no"]]) for e in estudiantes if e["no"] in FALTAS})
